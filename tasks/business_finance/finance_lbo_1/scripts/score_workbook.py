"""Local workbook scorer for business_finance/finance_lbo_1."""

from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


@dataclass
class WorkbookScoreResult:
    score: float
    passed: bool
    passed_checks: int
    total_checks: int
    reasons: list[str]


def parse_numeric(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped.startswith("=") or stripped.startswith("#"):
            return None
        cleaned = stripped.replace(",", "").replace("$", "").replace("%", "")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def score_workbook_bytes(
    agent_bytes: bytes,
    ground_truth: dict[str, Any],
    pass_threshold: int = 24,
) -> WorkbookScoreResult:
    try:
        workbook = load_workbook(BytesIO(agent_bytes), data_only=True)
    except Exception as exc:
        return WorkbookScoreResult(
            score=0.0,
            passed=False,
            passed_checks=0,
            total_checks=len(ground_truth),
            reasons=[f"unreadable_workbook:{exc}"],
        )

    reasons: list[str] = []
    passed_checks = 0
    total_checks = len(ground_truth)

    for coord, meta in ground_truth.items():
        sheet_name, cell_ref = coord.split("!")
        if sheet_name not in workbook.sheetnames:
            reasons.append(f"missing_sheet:{sheet_name}")
            continue
        submitted = workbook[sheet_name][cell_ref].value
        numeric = parse_numeric(submitted)
        if numeric is None:
            reasons.append(f"non_numeric_or_empty:{coord}")
            continue

        expected = float(meta["expected_value"])
        tolerance_pct = float(meta["tolerance_pct"])
        if expected == 0:
            ok = abs(numeric) < 1e-6
        else:
            ok = abs(numeric - expected) / abs(expected) <= tolerance_pct
        if ok:
            passed_checks += 1
        else:
            reasons.append(f"out_of_tolerance:{coord}")

    score = passed_checks / total_checks if total_checks else 0.0
    passed = passed_checks >= pass_threshold
    return WorkbookScoreResult(
        score=score,
        passed=passed,
        passed_checks=passed_checks,
        total_checks=total_checks,
        reasons=reasons,
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent", required=True)
    parser.add_argument("--ground-truth", required=True)
    parser.add_argument("--pass-threshold", type=int, default=24)
    args = parser.parse_args()

    with open(args.agent, "rb") as f:
        agent_bytes = f.read()
    with open(args.ground_truth, "r", encoding="utf-8") as f:
        ground_truth = json.load(f)

    result = score_workbook_bytes(
        agent_bytes=agent_bytes,
        ground_truth=ground_truth,
        pass_threshold=args.pass_threshold,
    )
    print(json.dumps(asdict(result), indent=2))


if __name__ == "__main__":
    main()
