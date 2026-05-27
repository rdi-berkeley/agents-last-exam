#!/usr/bin/env python
"""Score Mastercard income-statement workbook outputs."""

from __future__ import annotations

import argparse
import json
import math
import re
import warnings
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


PERIODS = ("2018A", "2019A", "2020E", "2021E", "2022E")
REQUIRED_PERIODS = ("2018A", "2019A", "2020Q3A", "2020E", "2021E", "2022E")
Q3_2020_ACTUAL_HEADERS = ("Mar-A", "Jun-A", "Sep-A")
FORECAST_PERIODS = {"2020E", "2021E", "2022E"}
REQUIRED_LABELS = (
    ("gross revenues", ("gross revenues", "total gross revenues")),
    ("rebates and incentives", ("rebates & incentives", "rebates and incentives")),
    ("net revenues", ("net revenues", "net revenue")),
    ("general and administrative", ("general and administrative",)),
    ("advertising and marketing", ("advertising and marketing",)),
    ("operating income", ("operating income",)),
    ("pretax income", ("pretax income", "pre-tax income", "income before taxes")),
    ("tax expense", ("income tax expense", "tax expense", "effective tax rate")),
    ("net income", ("net income", "net income to common shares")),
    ("basic eps", ("basic eps", "basic earnings per share")),
    ("diluted eps", ("diluted eps", "diluted earnings per share")),
    ("yoy growth", ("yoy growth", "year over year", "year-over-year")),
    ("operating margin", ("operating income margin", "operating margin")),
)


@dataclass(frozen=True)
class Metric:
    name: str
    aliases: tuple[str, ...]
    periods: tuple[str, ...] = PERIODS
    kind: str = "amount"


METRICS = (
    Metric("gross revenues", ("total gross revenues",)),
    Metric("net revenues", ("net revenues", "net revenue")),
    Metric("general and administrative", ("general and administrative",)),
    Metric("advertising and marketing", ("advertising and marketing",)),
    Metric("operating income", ("operating income",)),
    Metric("pretax income", ("pretax income", "pre-tax income", "income before taxes")),
    Metric("tax expense", ("income tax expense", "tax expense")),
    Metric("effective tax rate", ("effective tax rate",), kind="percent"),
    Metric("net income", ("net income", "net income to common shares")),
    Metric("diluted eps", ("diluted eps", "diluted earnings per share")),
    Metric("operating margin", ("operating income margin", "operating margin"), kind="percent"),
)


def _norm(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text.lower())
    return re.sub(r"\s+", " ", text).strip()


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value))


def _load_pair(workbook_bytes: bytes):
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    formulas = load_workbook(BytesIO(workbook_bytes), data_only=False, read_only=False)
    values = load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=False)
    return formulas, values


def _select_sheet(workbook) -> Any | None:
    if "IS" in workbook.sheetnames:
        return workbook["IS"]
    for sheet in workbook.worksheets:
        title = _norm(sheet.title)
        if "income statement" in title or title == "is":
            return sheet
    best_sheet = None
    best_hits = -1
    for sheet in workbook.worksheets:
        hits = 0
        for row in sheet.iter_rows():
            row_text = " ".join(_norm(cell.value) for cell in row[:8] if cell.value is not None)
            for _, aliases in REQUIRED_LABELS:
                if any(alias in row_text for alias in aliases):
                    hits += 1
                    break
        if hits > best_hits:
            best_sheet = sheet
            best_hits = hits
    return best_sheet if best_hits >= 3 else None


def _row_text(sheet, row_idx: int) -> str:
    return " ".join(_norm(sheet.cell(row=row_idx, column=col).value) for col in range(1, min(sheet.max_column, 10) + 1))


def _find_row(sheet, aliases: tuple[str, ...]) -> int | None:
    normalized_aliases = tuple(_norm(alias) for alias in aliases)
    exact_matches: list[int] = []
    fuzzy_matches: list[int] = []
    for row_idx in range(1, sheet.max_row + 1):
        row_values = [_norm(sheet.cell(row=row_idx, column=col).value) for col in range(1, min(sheet.max_column, 8) + 1)]
        if any(value in normalized_aliases for value in row_values if value):
            exact_matches.append(row_idx)
            continue
        text = " ".join(value for value in row_values if value)
        if any(alias and alias in text for alias in normalized_aliases):
            fuzzy_matches.append(row_idx)
    return (exact_matches or fuzzy_matches or [None])[0]


def _find_period_columns(sheet) -> dict[str, int]:
    columns: dict[str, int] = {}
    for period in PERIODS:
        period_norm = _norm(period)
        matches: list[int] = []
        for row_idx in range(1, min(sheet.max_row, 12) + 1):
            for col_idx in range(1, sheet.max_column + 1):
                if _norm(sheet.cell(row=row_idx, column=col_idx).value) == period_norm:
                    matches.append(col_idx)
        if matches:
            columns[period] = max(matches)
    return columns


def _find_2020q3_actual_columns(sheet) -> list[int]:
    expected = [_norm(value) for value in Q3_2020_ACTUAL_HEADERS]
    best: list[int] = []
    for row_idx in range(1, min(sheet.max_row, 12) + 1):
        values = [_norm(sheet.cell(row=row_idx, column=col_idx).value) for col_idx in range(1, sheet.max_column + 1)]
        for idx in range(0, len(values) - len(expected) + 1):
            if values[idx : idx + len(expected)] == expected:
                candidate = [idx + offset + 1 for offset in range(len(expected))]
                if not best or min(candidate) > min(best):
                    best = candidate
    return best


def _label_coverage(sheet) -> tuple[float, list[str]]:
    missing = []
    for label, aliases in REQUIRED_LABELS:
        if _find_row(sheet, aliases) is None:
            missing.append(label)
    return (len(REQUIRED_LABELS) - len(missing)) / len(REQUIRED_LABELS), missing


def _period_coverage(sheet, columns: dict[str, int]) -> tuple[float, list[str]]:
    missing = []
    for period in REQUIRED_PERIODS:
        if period == "2020Q3A":
            if len(_find_2020q3_actual_columns(sheet)) != len(Q3_2020_ACTUAL_HEADERS):
                missing.append(period)
        elif period not in columns:
            missing.append(period)
    return (len(REQUIRED_PERIODS) - len(missing)) / len(REQUIRED_PERIODS), missing


def _relative_ok(actual: float, expected: float, *, tolerance: float, abs_tolerance: float) -> bool:
    if abs(actual - expected) <= abs_tolerance:
        return True
    scale = max(abs(expected), 1.0)
    return abs(actual - expected) / scale <= tolerance


def _numeric_score(candidate_sheet, reference_sheet) -> dict[str, Any]:
    candidate_columns = _find_period_columns(candidate_sheet)
    reference_columns = _find_period_columns(reference_sheet)
    candidate_q3_columns = _find_2020q3_actual_columns(candidate_sheet)
    reference_q3_columns = _find_2020q3_actual_columns(reference_sheet)
    possible = 0
    passed = 0
    missing = []
    mismatches = []

    for metric in METRICS:
        candidate_row = _find_row(candidate_sheet, metric.aliases)
        reference_row = _find_row(reference_sheet, metric.aliases)
        if candidate_row is None or reference_row is None:
            missing.append(metric.name)
            possible += len(metric.periods)
            continue
        for period in metric.periods:
            possible += 1
            c_col = candidate_columns.get(period)
            r_col = reference_columns.get(period)
            if c_col is None or r_col is None:
                missing.append(f"{metric.name}:{period}")
                continue
            actual = candidate_sheet.cell(row=candidate_row, column=c_col).value
            expected = reference_sheet.cell(row=reference_row, column=r_col).value
            if not (_is_number(actual) and _is_number(expected)):
                missing.append(f"{metric.name}:{period}")
                continue
            actual_f = float(actual)
            expected_f = float(expected)
            if metric.kind == "percent":
                ok = abs(actual_f - expected_f) <= 0.05
            else:
                tolerance = 0.10 if period in FORECAST_PERIODS else 0.005
                ok = _relative_ok(actual_f, expected_f, tolerance=tolerance, abs_tolerance=1.0)
            if ok:
                passed += 1
            elif len(mismatches) < 20:
                mismatches.append(
                    {
                        "metric": metric.name,
                        "period": period,
                        "actual": actual_f,
                        "expected": expected_f,
                    }
                )
        if len(candidate_q3_columns) == len(Q3_2020_ACTUAL_HEADERS) and len(reference_q3_columns) == len(
            Q3_2020_ACTUAL_HEADERS
        ):
            for label, c_col, r_col in zip(Q3_2020_ACTUAL_HEADERS, candidate_q3_columns, reference_q3_columns):
                possible += 1
                actual = candidate_sheet.cell(row=candidate_row, column=c_col).value
                expected = reference_sheet.cell(row=reference_row, column=r_col).value
                if not (_is_number(actual) and _is_number(expected)):
                    missing.append(f"{metric.name}:2020Q3A:{label}")
                    continue
                actual_f = float(actual)
                expected_f = float(expected)
                ok = (
                    abs(actual_f - expected_f) <= 0.05
                    if metric.kind == "percent"
                    else _relative_ok(actual_f, expected_f, tolerance=0.005, abs_tolerance=1.0)
                )
                if ok:
                    passed += 1
                elif len(mismatches) < 20:
                    mismatches.append(
                        {
                            "metric": metric.name,
                            "period": f"2020Q3A:{label}",
                            "actual": actual_f,
                            "expected": expected_f,
                        }
                    )
        else:
            for label in Q3_2020_ACTUAL_HEADERS:
                possible += 1
                missing.append(f"{metric.name}:2020Q3A:{label}")

    score = passed / possible if possible else 0.0
    return {
        "score": score,
        "passed": passed,
        "possible": possible,
        "missing": missing[:40],
        "mismatches": mismatches,
    }


def _formula_score(candidate_formula_sheet, reference_formula_sheet, candidate_value_sheet) -> dict[str, Any]:
    reference_formula_count = sum(
        1
        for row in reference_formula_sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    candidate_formula_count = sum(
        1
        for row in candidate_formula_sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )

    candidate_columns = _find_period_columns(candidate_value_sheet)
    key_possible = 0
    key_formulas = 0
    for metric in METRICS:
        row_idx = _find_row(candidate_formula_sheet, metric.aliases)
        if row_idx is None:
            continue
        for period in metric.periods:
            col_idx = candidate_columns.get(period)
            if col_idx is None:
                continue
            key_possible += 1
            value = candidate_formula_sheet.cell(row=row_idx, column=col_idx).value
            if isinstance(value, str) and value.startswith("="):
                key_formulas += 1

    volume_score = min(candidate_formula_count / max(reference_formula_count * 0.20, 1), 1.0)
    key_score = key_formulas / key_possible if key_possible else 0.0
    return {
        "score": 0.45 * volume_score + 0.55 * key_score,
        "formula_cells": candidate_formula_count,
        "reference_formula_cells": reference_formula_count,
        "key_formula_cells": key_formulas,
        "key_formula_possible": key_possible,
    }


def score_workbooks(candidate_bytes: bytes, reference_bytes: bytes) -> dict[str, Any]:
    try:
        candidate_formulas, candidate_values = _load_pair(candidate_bytes)
        reference_formulas, reference_values = _load_pair(reference_bytes)
    except Exception as exc:
        return {"score": 0.0, "hard_fail": "unreadable_workbook", "error": str(exc)}

    candidate_formula_sheet = _select_sheet(candidate_formulas)
    candidate_value_sheet = _select_sheet(candidate_values)
    reference_formula_sheet = _select_sheet(reference_formulas)
    reference_value_sheet = _select_sheet(reference_values)
    if not all((candidate_formula_sheet, candidate_value_sheet, reference_formula_sheet, reference_value_sheet)):
        return {"score": 0.0, "hard_fail": "missing_income_statement_sheet"}

    label_score, missing_labels = _label_coverage(candidate_value_sheet)
    period_columns = _find_period_columns(candidate_value_sheet)
    period_score, missing_periods = _period_coverage(candidate_value_sheet, period_columns)
    if label_score < 0.45 or period_score < 0.60:
        return {
            "score": 0.0,
            "hard_fail": "insufficient_income_statement_structure",
            "label_coverage": label_score,
            "period_coverage": period_score,
            "missing_labels": missing_labels,
            "missing_periods": missing_periods,
        }

    numeric = _numeric_score(candidate_value_sheet, reference_value_sheet)
    formulas = _formula_score(candidate_formula_sheet, reference_formula_sheet, candidate_value_sheet)
    structure_score = 0.60 * label_score + 0.40 * period_score
    score = 0.25 * structure_score + 0.55 * numeric["score"] + 0.20 * formulas["score"]
    if score < 0.05:
        score = 0.0
    return {
        "score": max(0.0, min(1.0, score)),
        "structure_score": structure_score,
        "label_coverage": label_score,
        "period_coverage": period_score,
        "missing_labels": missing_labels,
        "missing_periods": missing_periods,
        "numeric": numeric,
        "formulas": formulas,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--candidate", required=True)
    parser.add_argument("--reference", required=True)
    args = parser.parse_args()
    with open(args.candidate, "rb") as candidate_file:
        candidate_bytes = candidate_file.read()
    with open(args.reference, "rb") as reference_file:
        reference_bytes = reference_file.read()
    print(json.dumps(score_workbooks(candidate_bytes, reference_bytes), indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
