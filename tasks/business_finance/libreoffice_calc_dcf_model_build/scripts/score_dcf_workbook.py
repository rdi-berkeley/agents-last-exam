"""Local scorer for business_finance/libreoffice_calc_dcf_model_build."""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from typing import Any

from openpyxl import load_workbook


def _round_whole(value: float) -> int:
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _round_price(value: float) -> float:
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _parse_csv_text(text: str) -> list[dict[str, float]]:
    rows: list[dict[str, float]] = []
    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        parsed: dict[str, float] = {}
        for key, value in row.items():
            if key == "fiscal_year":
                parsed[key] = int(value)
            else:
                parsed[key] = float(value)
        rows.append(parsed)
    return rows


@dataclass
class VisibleInputs:
    income_rows: list[dict[str, float]]
    balance_rows: list[dict[str, float]]
    cash_flow_rows: list[dict[str, float]]
    config: dict[str, Any]


@dataclass
class ScoreResult:
    score: float
    passed: bool
    reason: str
    details: dict[str, Any]


def _load_inputs(
    *,
    income_csv_text: str,
    balance_csv_text: str,
    cash_flow_csv_text: str,
    dcf_config_text: str,
) -> VisibleInputs:
    return VisibleInputs(
        income_rows=_parse_csv_text(income_csv_text),
        balance_rows=_parse_csv_text(balance_csv_text),
        cash_flow_rows=_parse_csv_text(cash_flow_csv_text),
        config=json.loads(dcf_config_text),
    )


def _compute_expected(inputs: VisibleInputs) -> dict[str, Any]:
    assumptions = inputs.config["assumptions"]
    last_revenue = inputs.income_rows[-1]["revenue"]
    growth_rates = assumptions["revenue_growth_rates"]
    projected_revenue: list[int] = []
    prev_revenue = last_revenue
    for growth in growth_rates:
        next_revenue = _round_whole(prev_revenue * (1 + growth))
        projected_revenue.append(next_revenue)
        prev_revenue = next_revenue

    ebitda = [_round_whole(value * assumptions["ebitda_margin"]) for value in projected_revenue]
    da = [_round_whole(value * assumptions["depreciation_rate_of_revenue"]) for value in projected_revenue]
    capex = [_round_whole(value * assumptions["capex_rate_of_revenue"]) for value in projected_revenue]
    nwc = [_round_whole(value * assumptions["nwc_change_rate_of_revenue"]) for value in projected_revenue]
    ufcf = [
        _round_whole(((ebitda[idx] - da[idx]) * (1 - assumptions["tax_rate"])) + da[idx] - capex[idx] - nwc[idx])
        for idx in range(len(projected_revenue))
    ]

    wacc = assumptions["wacc"]
    tgr = assumptions["terminal_growth_rate"]
    pv_fcf = [_round_whole(value / ((1 + wacc) ** (idx + 1))) for idx, value in enumerate(ufcf)]
    sum_pv = _round_whole(sum(pv_fcf))
    terminal_value = _round_whole(ufcf[-1] * (1 + tgr) / (wacc - tgr))
    pv_terminal_value = _round_whole(terminal_value / ((1 + wacc) ** len(ufcf)))
    enterprise_value = _round_whole(sum_pv + pv_terminal_value)
    net_debt = inputs.config["net_debt_usd_m"]
    equity_value = _round_whole(enterprise_value - net_debt)
    diluted_shares = inputs.config["shares_diluted_m"]
    implied_share_price = _round_price(equity_value / diluted_shares)

    sensitivity_grid: list[list[int]] = []
    for grid_wacc in inputs.config["sensitivity_axes"]["wacc_range"]:
        row: list[int] = []
        for grid_tgr in inputs.config["sensitivity_axes"]["tgr_range"]:
            pv_sum = sum(
                _round_whole(value / ((1 + grid_wacc) ** (idx + 1)))
                for idx, value in enumerate(ufcf)
            )
            rounded_terminal = _round_whole(ufcf[-1] * (1 + grid_tgr) / (grid_wacc - grid_tgr))
            rounded_pv_terminal = _round_whole(rounded_terminal / ((1 + grid_wacc) ** len(ufcf)))
            row.append(_round_whole(pv_sum + rounded_pv_terminal))
        sensitivity_grid.append(row)

    historical_rows = {
        "Revenue": [row["revenue"] for row in inputs.income_rows],
        "EBITDA": [row["ebitda"] for row in inputs.income_rows],
        "D&A": [row["depreciation_amortization"] for row in inputs.income_rows],
        "CapEx": [abs(row["capex"]) for row in inputs.balance_rows],
        "Net Income": [row["net_income"] for row in inputs.income_rows],
        "Free Cash Flow": [row["fcf"] for row in inputs.cash_flow_rows],
    }

    return {
        "historical_rows": historical_rows,
        "growth_rates": growth_rates,
        "projected_revenue": projected_revenue,
        "ebitda": ebitda,
        "da": da,
        "capex": capex,
        "nwc": nwc,
        "ufcf": ufcf,
        "pv_fcf": pv_fcf,
        "sum_pv": sum_pv,
        "terminal_value": terminal_value,
        "pv_terminal_value": pv_terminal_value,
        "enterprise_value": enterprise_value,
        "net_debt": net_debt,
        "equity_value": equity_value,
        "diluted_shares": diluted_shares,
        "implied_share_price": implied_share_price,
        "sensitivity_grid": sensitivity_grid,
    }


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _numeric_match(actual: Any, expected: float, *, tolerance: float = 1e-6) -> bool:
    if not isinstance(actual, (int, float)):
        return False
    return abs(float(actual) - float(expected)) <= tolerance


def _sheet_value_grid(ws, cells: list[str]) -> list[Any]:
    return [ws[cell].value for cell in cells]


def _year_matches(actual: Any, expected_year: int) -> bool:
    if isinstance(actual, (int, float)):
        return int(actual) == expected_year
    return str(actual).strip() == str(expected_year)


def score_workbook_bytes(
    *,
    workbook_bytes: bytes,
    income_csv_text: str | None = None,
    balance_csv_text: str | None = None,
    cash_flow_csv_text: str | None = None,
    dcf_config_text: str | None = None,
    output_contract_text: str | None = None,
    evaluation_context_text: str | None = None,
) -> ScoreResult:
    try:
        formula_wb = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
        value_wb = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    except Exception as exc:
        return ScoreResult(0.0, False, f"workbook is not a readable .xlsx file: {exc}", {})

    if evaluation_context_text:
        evaluation_context = json.loads(evaluation_context_text)
        snapshot = evaluation_context["visible_inputs_snapshot"]
        inputs = VisibleInputs(
            income_rows=snapshot["income_rows"],
            balance_rows=snapshot["balance_rows"],
            cash_flow_rows=snapshot["cash_flow_rows"],
            config=snapshot["config"],
        )
        contract = evaluation_context["output_contract"]
    else:
        if None in {
            income_csv_text,
            balance_csv_text,
            cash_flow_csv_text,
            dcf_config_text,
            output_contract_text,
        }:
            raise ValueError("either evaluation_context_text or all visible input texts must be provided")
        inputs = _load_inputs(
            income_csv_text=income_csv_text,
            balance_csv_text=balance_csv_text,
            cash_flow_csv_text=cash_flow_csv_text,
            dcf_config_text=dcf_config_text,
        )
        contract = json.loads(output_contract_text)
    expected = _compute_expected(inputs)
    layout = contract["sheet_layout"]

    required_sheets = contract["required_sheets"]
    if formula_wb.sheetnames != required_sheets:
        return ScoreResult(
            0.0,
            False,
            f"sheet names/order mismatch: expected {required_sheets}, got {formula_wb.sheetnames}",
            {"sheetnames": formula_wb.sheetnames},
        )

    for sheet_name in ("Projections", "DCF"):
        ws = formula_wb[sheet_name]
        for cell in contract["formula_cells"][sheet_name]:
            if not _is_formula(ws[cell].value):
                return ScoreResult(
                    0.0,
                    False,
                    f"required formula missing at {sheet_name}!{cell}",
                    {"sheet": sheet_name, "cell": cell, "value": ws[cell].value},
                )

    for cell in contract["formula_cells"]["Sensitivity"]:
        formula_value = formula_wb["Sensitivity"][cell].value
        cached_value = value_wb["Sensitivity"][cell].value
        if not _is_formula(formula_value) and not isinstance(cached_value, (int, float)):
            return ScoreResult(
                0.0,
                False,
                f"sensitivity cell {cell} is neither formula-backed nor numeric",
                {"cell": cell, "formula_value": formula_value, "cached_value": cached_value},
            )

    historical_ws = value_wb["Historical"]
    historical_years = [int(row["fiscal_year"]) for row in inputs.income_rows]
    actual_historical_years = _sheet_value_grid(historical_ws, layout["Historical"]["year_cells"])
    if any(not _year_matches(actual, expected_year) for actual, expected_year in zip(actual_historical_years, historical_years)):
        return ScoreResult(0.0, False, "historical year headers mismatch", {"actual": actual_historical_years, "expected": historical_years})
    for label, cells in layout["Historical"]["metric_rows"].items():
        row_idx = cells[0][1:]
        if historical_ws[f"A{row_idx}"].value != label:
            return ScoreResult(0.0, False, f"historical row label mismatch for {label}", {"actual": historical_ws[f'A{row_idx}'].value})
        row_values = _sheet_value_grid(historical_ws, cells)
        expected_values = expected["historical_rows"][label]
        if any(not _numeric_match(actual, target) for actual, target in zip(row_values, expected_values)):
            return ScoreResult(
                0.0,
                False,
                f"historical row mismatch for {label}",
                {"actual": row_values, "expected": expected_values},
            )

    projections_ws = value_wb["Projections"]
    projected_years = [inputs.config["base_year"] + offset for offset in range(1, inputs.config["projection_years"] + 1)]
    actual_projected_years = _sheet_value_grid(projections_ws, layout["Projections"]["year_cells"])
    if any(not _year_matches(actual, expected_year) for actual, expected_year in zip(actual_projected_years, projected_years)):
        return ScoreResult(0.0, False, "projection year headers mismatch", {"actual": actual_projected_years, "expected": projected_years})
    growth_cells = layout["Projections"]["growth_cells"]
    growth_values = _sheet_value_grid(projections_ws, growth_cells)
    if any(not _numeric_match(actual, target) for actual, target in zip(growth_values, expected["growth_rates"])):
        return ScoreResult(0.0, False, "projection growth rates do not match visible config", {"actual": growth_values})

    projection_checks = {
        "Revenue": expected["projected_revenue"],
        "EBITDA": expected["ebitda"],
        "D&A": expected["da"],
        "CapEx": expected["capex"],
        "Change in NWC": expected["nwc"],
        "Unlevered FCF": expected["ufcf"],
    }
    for label, expected_values in projection_checks.items():
        row_idx = layout["Projections"]["metric_rows"][label][0][1:]
        if projections_ws[f"A{row_idx}"].value != label:
            return ScoreResult(0.0, False, f"projection row label mismatch for {label}", {"actual": projections_ws[f'A{row_idx}'].value})
        actual_values = _sheet_value_grid(projections_ws, layout["Projections"]["metric_rows"][label])
        if any(not _numeric_match(actual, target) for actual, target in zip(actual_values, expected_values)):
            return ScoreResult(
                0.0,
                False,
                f"projection row mismatch for {label}",
                {"actual": actual_values, "expected": expected_values},
            )

    dcf_ws = value_wb["DCF"]
    assumptions = inputs.config["assumptions"]
    dcf_assumption_checks = {
        layout["DCF"]["assumption_cells"]["WACC"]: assumptions["wacc"],
        layout["DCF"]["assumption_cells"]["Terminal Growth Rate"]: assumptions["terminal_growth_rate"],
        layout["DCF"]["assumption_cells"]["Tax Rate"]: assumptions["tax_rate"],
        layout["DCF"]["assumption_cells"]["EBITDA Margin"]: assumptions["ebitda_margin"],
        layout["DCF"]["assumption_cells"]["D&A / Revenue"]: assumptions["depreciation_rate_of_revenue"],
        layout["DCF"]["assumption_cells"]["CapEx / Revenue"]: assumptions["capex_rate_of_revenue"],
        layout["DCF"]["assumption_cells"]["Change in NWC / Revenue"]: assumptions["nwc_change_rate_of_revenue"],
    }
    for cell, target in dcf_assumption_checks.items():
        row_idx = cell[1:]
        expected_label = next(label for label, label_cell in layout["DCF"]["assumption_cells"].items() if label_cell == cell)
        if dcf_ws[f"A{row_idx}"].value != expected_label:
            return ScoreResult(0.0, False, f"DCF label mismatch for {expected_label}", {"actual": dcf_ws[f'A{row_idx}'].value})
        if not _numeric_match(dcf_ws[cell].value, target):
            return ScoreResult(
                0.0,
                False,
                f"DCF assumption mismatch at {cell}",
                {"actual": dcf_ws[cell].value, "expected": target},
            )

    dcf_checks = {
        tuple(layout["DCF"]["pv_fcf_cells"]): expected["pv_fcf"],
        layout["DCF"]["summary_cells"]["Sum PV FCFs ($M)"]: expected["sum_pv"],
        layout["DCF"]["summary_cells"]["Terminal Value ($M)"]: expected["terminal_value"],
        layout["DCF"]["summary_cells"]["PV Terminal Value ($M)"]: expected["pv_terminal_value"],
        layout["DCF"]["summary_cells"]["Enterprise Value ($M)"]: expected["enterprise_value"],
        layout["DCF"]["summary_cells"]["Net Debt ($M)"]: expected["net_debt"],
        layout["DCF"]["summary_cells"]["Equity Value ($M)"]: expected["equity_value"],
        layout["DCF"]["summary_cells"]["Diluted Shares (M)"]: expected["diluted_shares"],
    }
    for key, target in dcf_checks.items():
        if isinstance(key, tuple):
            for label, label_cell in layout["DCF"]["pv_fcf_label_cells"].items():
                if dcf_ws[label_cell].value != label:
                    return ScoreResult(
                        0.0,
                        False,
                        "DCF PV row labels are not in the published layout",
                        {"cell": label_cell, "actual": dcf_ws[label_cell].value, "expected": label},
                    )
            actual_values = [dcf_ws[cell].value for cell in key]
            if any(not _numeric_match(actual, expected_value) for actual, expected_value in zip(actual_values, target)):
                return ScoreResult(
                    0.0,
                    False,
                    "DCF PV FCF row mismatch",
                    {"actual": actual_values, "expected": target},
                )
        else:
            row_idx = key[1:]
            expected_label = next(label for label, label_cell in layout["DCF"]["summary_cells"].items() if label_cell == key)
            if dcf_ws[f"A{row_idx}"].value != expected_label:
                return ScoreResult(0.0, False, f"DCF summary label mismatch for {expected_label}", {"actual": dcf_ws[f'A{row_idx}'].value})
            if not _numeric_match(dcf_ws[key].value, target):
                return ScoreResult(
                    0.0,
                    False,
                    f"DCF mismatch at {key}",
                    {"actual": dcf_ws[key].value, "expected": target},
                )

    implied_share_price_cell = layout["DCF"]["summary_cells"]["Implied Share Price ($)"]
    if not _numeric_match(dcf_ws[implied_share_price_cell].value, expected["implied_share_price"], tolerance=0.01):
        return ScoreResult(
            0.0,
            False,
            "implied share price mismatch",
            {"actual": dcf_ws[implied_share_price_cell].value, "expected": expected["implied_share_price"]},
        )

    sensitivity_ws = value_wb["Sensitivity"]
    header_cell = layout["Sensitivity"]["top_left_header_cell"]
    if sensitivity_ws[header_cell].value != contract["sensitivity_grid"]["top_left_header"]:
        return ScoreResult(0.0, False, "sensitivity top-left header mismatch", {"actual": sensitivity_ws[header_cell].value})

    row_labels = _sheet_value_grid(sensitivity_ws, layout["Sensitivity"]["row_label_cells"])
    col_labels = _sheet_value_grid(sensitivity_ws, layout["Sensitivity"]["column_label_cells"])
    if any(not _numeric_match(actual, target) for actual, target in zip(row_labels, contract["sensitivity_grid"]["row_labels"])):
        return ScoreResult(0.0, False, "sensitivity row labels mismatch", {"actual": row_labels})
    if any(not _numeric_match(actual, target) for actual, target in zip(col_labels, contract["sensitivity_grid"]["column_labels"])):
        return ScoreResult(0.0, False, "sensitivity column labels mismatch", {"actual": col_labels})

    for row_idx, expected_row in enumerate(expected["sensitivity_grid"]):
        actual_row = _sheet_value_grid(sensitivity_ws, layout["Sensitivity"]["grid_cells"][row_idx])
        if any(not _numeric_match(actual, expected_value) for actual, expected_value in zip(actual_row, expected_row)):
            return ScoreResult(
                0.0,
                False,
                f"sensitivity row mismatch at row {row_idx + 3}",
                {"actual": actual_row, "expected": expected_row},
            )

    return ScoreResult(
        1.0,
        True,
        "workbook matches the visible DCF contract",
        {
            "enterprise_value": expected["enterprise_value"],
            "implied_share_price": expected["implied_share_price"],
            "sensitivity_base_case": expected["sensitivity_grid"][2][2],
        },
    )
