"""Layout-agnostic workbook scorer for business_finance/equity_research_summary.

Finds labels and section headers by scanning the worksheet instead of
requiring fixed cell coordinates. This allows the agent to use any layout.
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import asdict, dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


@dataclass
class WorkbookScoreResult:
    score: float
    passed: bool
    reasons: list[str]


def _parse_numeric(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped.startswith("="):
            return None
        cleaned = stripped.replace(",", "").replace("$", "").replace("%", "")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _has_fill(cell) -> bool:
    fill = cell.fill
    if not fill.fill_type:
        return False
    rgb = (fill.fgColor.rgb or "").upper()
    return rgb not in {"", "00000000", "00FFFFFF", "FFFFFFFF"}


def _normalize(text: str) -> str:
    return re.sub(r"\s+", " ", str(text).strip().lower())


def _extract_cell_refs(formula: str) -> set[str]:
    return set(re.findall(r"\b([A-Z]+[0-9]+)\b", formula.upper()))


def _build_label_map(ws) -> dict[str, tuple[str, Any]]:
    """Map normalized label text -> (value_cell_coordinate, value_cell_object).

    A label is any non-formula text cell. Its value cell is the cell in the
    next column of the same row.
    """
    label_map: dict[str, tuple[str, Any]] = {}
    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            if (
                cell.value
                and isinstance(cell.value, str)
                and not cell.value.startswith("=")
            ):
                norm = _normalize(cell.value)
                if norm in label_map:
                    continue
                if i + 1 < len(row):
                    val_cell = row[i + 1]
                    label_map[norm] = (val_cell.coordinate, val_cell)
    return label_map


def _find_label(
    label_map: dict[str, tuple[str, Any]], target: str
) -> tuple[str, Any] | None:
    norm_target = _normalize(target)
    if norm_target in label_map:
        return label_map[norm_target]
    for key, val in label_map.items():
        if key.startswith(norm_target) or norm_target in key:
            return val
    return None


def score_workbook_bytes(
    agent_bytes: bytes, manifest: dict[str, Any]
) -> WorkbookScoreResult:
    try:
        wb = load_workbook(BytesIO(agent_bytes), data_only=False)
    except Exception as exc:
        return WorkbookScoreResult(
            score=0.0, passed=False, reasons=[f"unreadable_workbook:{exc}"]
        )

    ws = wb.active
    reasons: list[str] = []
    total = 0
    passed = 0

    sheet_check = manifest.get("sheet_name_contains", "")
    total += 1
    if sheet_check.lower() in (ws.title or "").lower():
        passed += 1
    else:
        reasons.append(f"sheet_name_missing:{sheet_check}")

    label_map = _build_label_map(ws)

    for header in manifest.get("section_headers", []):
        total += 1
        text = header["text"]
        norm_text = _normalize(text)
        found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if norm_text in _normalize(cell.value):
                        found = True
                        ok = True
                        if header.get("bold") and not (cell.font and cell.font.bold):
                            reasons.append(f"header_not_bold:{text}")
                            ok = False
                        if header.get("fill") and not _has_fill(cell):
                            reasons.append(f"header_no_fill:{text}")
                            ok = False
                        if ok:
                            passed += 1
                        break
            if found:
                break
        if not found:
            reasons.append(f"header_missing:{text}")

    for entry in manifest.get("fixed_values", []):
        total += 1
        label = entry["label"]
        expected = entry["value"]
        tolerance = entry.get("tolerance", 1.0)
        result = _find_label(label_map, label)
        if result is None:
            reasons.append(f"fixed_label_missing:{label}")
            continue
        _, cell = result
        actual = _parse_numeric(cell.value)
        if actual is None:
            reasons.append(f"fixed_value_not_numeric:{label}")
            continue
        if abs(actual - expected) > tolerance:
            reasons.append(f"fixed_value_mismatch:{label}:{actual}!={expected}")
        else:
            passed += 1

    for entry in manifest.get("formula_cells", []) + manifest.get(
        "dashboard_formulas", []
    ):
        total += 1
        label = entry["label"]
        must_ref_labels = entry["must_reference"]
        result = _find_label(label_map, label)
        if result is None:
            reasons.append(f"formula_label_missing:{label}")
            continue
        _, cell = result
        if not isinstance(cell.value, str) or not cell.value.startswith("="):
            reasons.append(f"not_a_formula:{label}")
            continue
        formula_refs = _extract_cell_refs(cell.value)
        all_ok = True
        for ref_label in must_ref_labels:
            ref_result = _find_label(label_map, ref_label)
            if ref_result is None:
                all_ok = False
                reasons.append(f"formula_ref_label_missing:{label}->{ref_label}")
                break
            ref_coord = ref_result[0]
            if ref_coord.upper() not in formula_refs:
                all_ok = False
                reasons.append(
                    f"formula_missing_ref:{label}->{ref_label}({ref_coord})"
                )
                break
        if all_ok:
            passed += 1

    for label in manifest.get("live_labels", []):
        total += 1
        result = _find_label(label_map, label)
        if result is None:
            reasons.append(f"live_label_missing:{label}")
            continue
        _, cell = result
        actual = _parse_numeric(cell.value)
        if actual is None:
            reasons.append(f"live_value_not_numeric:{label}")
        else:
            passed += 1

    score = round(passed / total, 4) if total > 0 else 0.0
    return WorkbookScoreResult(score=score, passed=(score >= 1.0), reasons=reasons)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent", required=True)
    parser.add_argument("--manifest", required=True)
    args = parser.parse_args()

    with open(args.agent, "rb") as f:
        agent_bytes = f.read()
    with open(args.manifest, "r", encoding="utf-8") as f:
        manifest = json.load(f)

    result = score_workbook_bytes(agent_bytes=agent_bytes, manifest=manifest)
    print(json.dumps(asdict(result), indent=2))


if __name__ == "__main__":
    main()
