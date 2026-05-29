"""Score insurance settlement report workbooks."""

from __future__ import annotations

import argparse
import json
import math
import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

EXPECTED_HEADERS = [
    "Claim_Ref",
    "Policy_Number",
    "Insured_Name_Extracted",
    "Insured_Name_Standardized",
    "Date_of_Loss",
    "Peril",
    "Claimed_Amount",
    "Deductible_Applied",
    "Deductible_Basis",
    "SubLimit_Applied",
    "Coverage_Ratio",
    "Net_Loss_After_Deductible",
    "Final_Payable_Amount",
    "Data_Cleaning_Notes",
    "Amount_Currency",
]

FIELD_COLUMNS = [
    "Claim_Ref",
    "Policy_Number",
    "Insured_Name_Standardized",
    "Date_of_Loss",
    "Peril",
]

EXPECTED_CLAIM_COUNT = 7
FIELD_PASS_THRESHOLD = 33
FINAL_PAYABLE_TOLERANCE = 1.0
TOTAL_PAYABLE_TOLERANCE = 5.0
NOTE_PASS_THRESHOLD = 3

ISSUE_ROWS = {
    "clm002_missing_deductible": "CLM-2024-0315-002",
    "clm002_name_discrepancy": "CLM-2024-0315-002",
    "clm004_same_as_main_policy": "CLM-2024-0322-004",
    "clm005_name_abbreviation": "CLM-2024-0325-005",
}

NEGATION_TERMS = r"\b(?:no|not|none|without|never|neither|nor)\b"


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return isinstance(value, str) and value.strip() == ""


def _canon_text(value: Any) -> str:
    if _is_blank(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _canon_date(value: Any) -> str:
    if _is_blank(value):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _canon_text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def _as_number(value: Any) -> float | None:
    if _is_blank(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value)
    text = text.replace(",", "").replace("RMB", "").replace("CNY", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def _load_claim_rows(workbook_bytes: bytes) -> tuple[dict[str, dict[str, Any]], list[str]]:
    wb = load_workbook(BytesIO(workbook_bytes), data_only=True)
    ws = wb["ClaimsSettlement"] if "ClaimsSettlement" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("workbook contains no rows")

    headers = [_canon_text(value) for value in rows[0][: len(EXPECTED_HEADERS)]]
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"unexpected header: {headers}")

    claim_rows: dict[str, dict[str, Any]] = {}
    for raw_row in rows[1:]:
        record = dict(zip(EXPECTED_HEADERS, raw_row[: len(EXPECTED_HEADERS)]))
        claim_ref = _canon_text(record.get("Claim_Ref"))
        if not claim_ref or claim_ref.upper() == "TOTAL":
            continue
        claim_rows[claim_ref] = record
    return claim_rows, headers


def _same_field(candidate: Any, reference: Any, column: str) -> bool:
    if column == "Date_of_Loss":
        return _canon_date(candidate) == _canon_date(reference)
    return _canon_text(candidate) == _canon_text(reference)


def _contains(pattern: str, text: str) -> bool:
    return re.search(pattern, text, flags=re.IGNORECASE) is not None


def _negated_near(pattern: str, text: str, *, window: int = 40) -> bool:
    for match in re.finditer(pattern, text, flags=re.IGNORECASE):
        start = max(0, match.start() - window)
        end = min(len(text), match.end() + window)
        if _contains(NEGATION_TERMS, text[start:end]):
            return True
    return False


def _issue_present(issue: str, note: Any) -> bool:
    text = _canon_text(note).lower()
    if not text:
        return False

    if issue == "clm002_missing_deductible":
        if _contains(
            r"\bno\s+(?:missing|blank|absent|omitted)?\s*deductible\b|"
            r"\bdeductible\s+(?:was\s+)?not\s+(?:missing|blank|absent|omitted)\b|"
            r"\bwithout\s+(?:a\s+)?(?:missing|blank|absent|omitted)\s+deductible\b",
            text,
        ):
            return False
        return (
            _contains(r"deductible", text)
            and _contains(r"missing|not provided|blank|absent|omitted|not stated|not shown", text)
            and _contains(r"policy|terms|lookup|looked up|retriev|resolved|cross-reference", text)
        )

    if issue == "clm002_name_discrepancy":
        if _negated_near(r"discrep|alias|parenthes|chinese|mismatch|different", text):
            return False
        return (
            _contains(r"name|insured", text)
            and _contains(r"discrep|alias|parenthes|chinese|mismatch|different", text)
            and _contains(r"standardi[sz]ed|policy record|policy name|resolved", text)
        )

    if issue == "clm004_same_as_main_policy":
        if _negated_near(r"same\s+as\s+main\s+policy|ambiguous|main policy", text):
            return False
        return (
            _contains(r"same\s+as\s+main\s+policy|ambiguous|main policy", text)
            and _contains(r"deductible", text)
            and _contains(r"lookup|looked up|resolved|cross-reference|policy terms|applied", text)
        )

    if issue == "clm005_name_abbreviation":
        if _negated_near(r"abbreviat|short|truncated", text):
            return False
        return (
            _contains(r"name|insured", text)
            and _contains(r"abbreviat|short|truncated", text)
            and _contains(r"standardi[sz]ed|policy record|company|co\.", text)
        )

    raise ValueError(f"unknown issue: {issue}")


def score_workbooks(candidate_bytes: bytes, reference_bytes: bytes) -> dict[str, Any]:
    try:
        candidate_rows, _ = _load_claim_rows(candidate_bytes)
        reference_rows, _ = _load_claim_rows(reference_bytes)
    except Exception as exc:
        return {"score": 0.0, "passed": False, "error": str(exc)}

    expected_refs = list(reference_rows)
    missing_claims = [claim_ref for claim_ref in expected_refs if claim_ref not in candidate_rows]
    extra_claims = [claim_ref for claim_ref in candidate_rows if claim_ref not in reference_rows]
    if missing_claims or len(candidate_rows) < EXPECTED_CLAIM_COUNT:
        return {
            "score": 0.0,
            "passed": False,
            "missing_claims": missing_claims,
            "extra_claims": extra_claims,
        }

    field_correct = 0
    field_mismatches = []
    final_amount_errors = {}
    final_amounts_pass = True
    candidate_total = 0.0
    reference_total = 0.0

    for claim_ref in expected_refs:
        candidate = candidate_rows[claim_ref]
        reference = reference_rows[claim_ref]
        for column in FIELD_COLUMNS:
            if _same_field(candidate.get(column), reference.get(column), column):
                field_correct += 1
            else:
                field_mismatches.append(
                    {
                        "claim_ref": claim_ref,
                        "column": column,
                        "candidate": _canon_text(candidate.get(column)),
                        "reference": _canon_text(reference.get(column)),
                    }
                )

        candidate_amount = _as_number(candidate.get("Final_Payable_Amount"))
        reference_amount = _as_number(reference.get("Final_Payable_Amount"))
        if candidate_amount is None or reference_amount is None:
            final_amount_errors[claim_ref] = None
            final_amounts_pass = False
            continue
        error = abs(candidate_amount - reference_amount)
        final_amount_errors[claim_ref] = error
        candidate_total += candidate_amount
        reference_total += reference_amount
        if error > FINAL_PAYABLE_TOLERANCE:
            final_amounts_pass = False

    issue_hits = {}
    for issue, claim_ref in ISSUE_ROWS.items():
        issue_hits[issue] = _issue_present(issue, candidate_rows.get(claim_ref, {}).get("Data_Cleaning_Notes"))

    total_error = abs(candidate_total - reference_total)
    field_pass = field_correct >= FIELD_PASS_THRESHOLD
    total_pass = total_error <= TOTAL_PAYABLE_TOLERANCE
    note_count = sum(1 for hit in issue_hits.values() if hit)
    notes_pass = note_count >= NOTE_PASS_THRESHOLD
    passed = field_pass and final_amounts_pass and total_pass and notes_pass

    return {
        "score": 1.0 if passed else 0.0,
        "passed": passed,
        "field_correct": field_correct,
        "field_total": len(expected_refs) * len(FIELD_COLUMNS),
        "field_pass": field_pass,
        "field_mismatches": field_mismatches[:20],
        "final_amount_errors": final_amount_errors,
        "final_amounts_pass": final_amounts_pass,
        "candidate_total": candidate_total,
        "reference_total": reference_total,
        "total_error": total_error,
        "total_pass": total_pass,
        "issue_hits": issue_hits,
        "issue_count": note_count,
        "notes_pass": notes_pass,
        "extra_claims": extra_claims,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--candidate", required=True)
    parser.add_argument("--reference", required=True)
    args = parser.parse_args()
    report = score_workbooks(Path(args.candidate).read_bytes(), Path(args.reference).read_bytes())
    print(json.dumps(report, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
