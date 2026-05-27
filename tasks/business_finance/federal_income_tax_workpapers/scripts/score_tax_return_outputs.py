"""Scorer for federal_income_tax_workpapers outputs."""

import argparse
import json
import re
from io import BytesIO
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from pypdf import PdfReader

REQUIRED_SHEETNAMES = [
    "Summary",
    "Income and Deductions",
    "Tax Liability",
    "Credits and Payments",
]

CELL_REF_PATTERN = re.compile(r"(?:'[^']+'!)?\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?")

WORKBOOK_CONCEPT_CHECKS = [
    {
        "sheet": "Summary",
        "check_id": "summary_wages",
        "concept": "wages",
        "aliases": ("w-2", "w2", "wages", "wages salaries tips"),
        "expected_value": 40129.0,
    },
    {
        "sheet": "Summary",
        "check_id": "summary_interest",
        "concept": "interest",
        "aliases": ("interest", "taxable interest"),
        "expected_value": 779.0,
    },
    {
        "sheet": "Summary",
        "check_id": "summary_agi",
        "concept": "agi",
        "aliases": ("agi", "adjusted gross income"),
        "expected_value": [40908.0, 40608.0],
        "require_formula": True,
        "min_formula_refs": 2,
    },
    {
        "sheet": "Summary",
        "check_id": "summary_standard_deduction",
        "concept": "standard_deduction",
        "aliases": ("standard deduction",),
        "expected_abs_value": [15750.0, 15000.0],
    },
    {
        "sheet": "Summary",
        "check_id": "summary_taxable_income",
        "concept": "taxable_income",
        "aliases": ("taxable income",),
        "expected_value": [25158.0, 25608.0],
        "require_formula": True,
        "min_formula_refs": 2,
    },
    {
        "sheet": "Summary",
        "check_id": "summary_total_tax",
        "concept": "total_tax",
        "aliases": ("total tax",),
        "expected_value": [2780.34, 2834.34],
        "require_formula": True,
        "min_formula_refs": 1,
    },
    {
        "sheet": "Summary",
        "check_id": "summary_withholding",
        "concept": "withholding",
        "aliases": (
            "income tax withheld",
            "federal income tax withheld",
            "tax withheld",
            "withholding",
        ),
        "expected_value": 4564.0,
    },
    {
        "sheet": "Summary",
        "check_id": "summary_total_payments",
        "concept": "total_payments",
        "aliases": ("total payments",),
        "expected_value": 4564.0,
        "require_formula": True,
        "min_formula_refs": 1,
    },
    {
        "sheet": "Summary",
        "check_id": "summary_refund_or_liability",
        "concept": "refund_or_liability",
        "aliases": ("tax liability refund", "tax liability (refund)", "refund", "amount owed"),
        "expected_abs_value": [1783.66, 1729.66],
        "require_formula": True,
        "min_formula_refs": 2,
    },
    {
        "sheet": "Income and Deductions",
        "check_id": "income_wages",
        "concept": "wages",
        "aliases": ("w-2", "w2", "wages", "wages salaries tips"),
        "expected_value": 40129.0,
    },
    {
        "sheet": "Income and Deductions",
        "check_id": "income_interest",
        "concept": "interest",
        "aliases": ("interest", "taxable interest"),
        "expected_value": 779.0,
    },
    {
        "sheet": "Income and Deductions",
        "check_id": "income_agi",
        "concept": "agi",
        "aliases": ("agi", "adjusted gross income"),
        "expected_value": [40908.0, 40608.0],
        "require_formula": True,
        "min_formula_refs": 2,
    },
    {
        "sheet": "Income and Deductions",
        "check_id": "income_standard_deduction",
        "concept": "standard_deduction",
        "aliases": ("standard deduction",),
        "expected_abs_value": [15750.0, 15000.0],
    },
    {
        "sheet": "Tax Liability",
        "check_id": "tax_sheet_taxable_income",
        "concept": "taxable_income",
        "aliases": ("taxable income",),
        "expected_value": [25158.0, 25608.0],
        "require_formula": True,
        "min_formula_refs": 1,
    },
    {
        "sheet": "Tax Liability",
        "check_id": "tax_sheet_total_tax",
        "concept": "total_tax",
        "aliases": ("total", "total tax", "tax liability total"),
        "expected_value": [2780.34, 2834.34],
        "require_formula": True,
        "min_formula_refs": 2,
    },
    {
        "sheet": "Credits and Payments",
        "check_id": "credits_withholding",
        "concept": "withholding",
        "aliases": (
            "income tax withheld",
            "federal income tax withheld",
            "tax withheld",
            "withholding",
        ),
        "expected_value": 4564.0,
    },
    {
        "sheet": "Credits and Payments",
        "check_id": "credits_total_payments",
        "concept": "total_payments",
        "aliases": ("total payments",),
        "expected_value": 4564.0,
        "require_formula": True,
        "min_formula_refs": 1,
    },
]

PDF_FIELD_EXPECTATIONS = {
    "f1_14[0]": "Kristen",
    "f1_15[0]": "DeMarco",
    "f1_20[0]": "1234 Swing Street",
    "f1_22[0]": "Austin",
    "f1_23[0]": "TX",
    "f1_24[0]": "78705",
    "f1_47[0]": "40,129",
    "f1_57[0]": "40,129",
    "f1_59[0]": "779",
    "f1_75[0]": ("40,908", "40,608"),
    "f2_01[0]": ("40,908", "40,608"),
    "f2_06[0]": ("25,158", "25,608"),
    "f2_08[0]": ("2,780", "2,834"),
    "f2_10[0]": ("2,780", "2,834"),
    "f2_14[0]": ("2,780", "2,834"),
    "f2_16[0]": ("2,780", "2,834"),
    "f2_17[0]": "4,564",
    "f2_20[0]": "4,564",
    "f2_29[0]": "4,564",
    "f2_30[0]": ("1,784", "1,730"),
}

PDF_RECONCILIATION_CHECKS = [
    {"concept": "wages", "suffixes": ("f1_47[0]", "f1_57[0]"), "formatter": "whole_dollars"},
    {"concept": "interest", "suffixes": ("f1_59[0]",), "formatter": "whole_dollars"},
    {"concept": "agi", "suffixes": ("f1_75[0]", "f2_01[0]"), "formatter": "whole_dollars"},
    {"concept": "taxable_income", "suffixes": ("f2_06[0]",), "formatter": "whole_dollars"},
    {
        "concept": "total_tax",
        "suffixes": ("f2_08[0]", "f2_10[0]", "f2_14[0]", "f2_16[0]"),
        "formatter": "whole_dollars",
    },
    {
        "concept": "total_payments",
        "suffixes": ("f2_17[0]", "f2_20[0]", "f2_29[0]"),
        "formatter": "whole_dollars",
    },
    {
        "concept": "refund_or_liability",
        "suffixes": ("f2_30[0]",),
        "formatter": "whole_dollars_abs",
    },
]


def _normalize_formula(formula: Any) -> str:
    if formula is None:
        return ""
    text = str(formula).strip()
    if text.startswith("="):
        text = text[1:]
    return text.replace(" ", "").replace("'", "").replace("$", "").upper()


def _normalize_label(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())


def _coerce_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return float(value)
    except Exception:
        return None


def _close_enough(actual: Any, expected: float, tolerance: float = 0.01) -> bool:
    actual_number = _coerce_float(actual)
    if actual_number is None:
        return False
    return abs(actual_number - expected) <= tolerance


def _close_enough_abs(actual: Any, expected_abs: float, tolerance: float = 0.01) -> bool:
    actual_number = _coerce_float(actual)
    if actual_number is None:
        return False
    return abs(abs(actual_number) - expected_abs) <= tolerance


def _count_formula_refs(formula: Any) -> int:
    if not isinstance(formula, str) or not formula.startswith("="):
        return 0
    ref_count = 0
    for match in CELL_REF_PATTERN.findall(formula):
        if ":" in match:
            ref_count += 2
        else:
            ref_count += 1
    return ref_count


def _format_whole_dollars(value: Any, *, absolute: bool = False) -> str:
    number = _coerce_float(value)
    if number is None:
        raise ValueError(f"Cannot format non-numeric value: {value!r}")
    if absolute:
        number = abs(number)
    return f"{int(round(number)):,}"


PDF_FORMATTERS: dict[str, Callable[[Any], str]] = {
    "whole_dollars": lambda value: _format_whole_dollars(value, absolute=False),
    "whole_dollars_abs": lambda value: _format_whole_dollars(value, absolute=True),
}


def _read_workbook_objects(workbook_bytes: bytes):
    workbook_formula = load_workbook(BytesIO(workbook_bytes), data_only=False)
    workbook_values = load_workbook(BytesIO(workbook_bytes), data_only=True)
    return workbook_formula, workbook_values


def _extract_pdf_values(pdf_bytes: bytes) -> dict[str, str]:
    reader = PdfReader(BytesIO(pdf_bytes))
    fields = reader.get_fields() or {}
    values: dict[str, str] = {}
    for field_name, field_meta in fields.items():
        raw_value = field_meta.get("/V")
        if raw_value is None:
            continue
        text_value = str(raw_value).strip()
        if text_value == "":
            continue
        values[field_name] = text_value
    return values


def _lookup_pdf_value(pdf_values: dict[str, str], suffix: str) -> str | None:
    if suffix in pdf_values:
        return pdf_values[suffix]

    matches = [value for name, value in pdf_values.items() if name.endswith(suffix)]
    if not matches:
        return None
    if len(set(matches)) == 1:
        return matches[0]
    return None


def _label_match_score(cell_label: str, alias: str) -> int:
    if not cell_label or not alias:
        return 0
    if cell_label == alias:
        return 3
    if alias in cell_label:
        return 2
    if cell_label in alias:
        return 1
    return 0


def _find_label_cell(sheet, aliases: tuple[str, ...]):
    normalized_aliases = tuple(_normalize_label(alias) for alias in aliases)
    best_match = None
    best_score = 0
    best_coord = None

    for row in sheet.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            cell_label = _normalize_label(cell.value)
            score = max((_label_match_score(cell_label, alias) for alias in normalized_aliases), default=0)
            if score == 0:
                continue
            coord_rank = (cell.row, cell.column)
            if score > best_score or (score == best_score and (best_coord is None or coord_rank < best_coord)):
                best_match = cell
                best_score = score
                best_coord = coord_rank

    return best_match


def _cell_has_candidate_value(formula_value: Any, cached_value: Any) -> bool:
    if isinstance(formula_value, str) and formula_value.startswith("="):
        return True
    return _coerce_float(cached_value) is not None or _coerce_float(formula_value) is not None


def _find_value_cell(formula_sheet, value_sheet, label_cell):
    candidates = []
    for row_offset in (0, 1, -1):
        row_index = label_cell.row + row_offset
        if row_index < 1 or row_index > formula_sheet.max_row:
            continue
        for column_index in range(1, formula_sheet.max_column + 1):
            if row_index == label_cell.row and column_index == label_cell.column:
                continue
            formula_cell = formula_sheet.cell(row=row_index, column=column_index)
            value_cell = value_sheet.cell(row=row_index, column=column_index)
            if not _cell_has_candidate_value(formula_cell.value, value_cell.value):
                continue

            row_distance = abs(row_index - label_cell.row)
            column_distance = abs(column_index - label_cell.column)
            side_rank = 0 if column_index > label_cell.column else 1
            formula_rank = 0 if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else 1
            candidates.append(
                (
                    row_distance,
                    column_distance,
                    side_rank,
                    formula_rank,
                    row_index,
                    column_index,
                    formula_cell,
                    value_cell,
                )
            )

    if not candidates:
        return None

    candidates.sort()
    return candidates[0][-2], candidates[0][-1]


def _concept_values_consistent(existing: Any, new_value: Any, *, absolute: bool = False) -> bool:
    if absolute:
        return _close_enough_abs(existing, abs(float(new_value)))
    return _close_enough(existing, float(new_value))


def _check_workbook(workbook_bytes: bytes) -> tuple[list[str], dict[str, Any], Any]:
    errors: list[str] = []
    workbook_formula, workbook_values = _read_workbook_objects(workbook_bytes)

    if workbook_formula.sheetnames != REQUIRED_SHEETNAMES:
        errors.append(
            "Workbook sheetnames must be exactly "
            f"{REQUIRED_SHEETNAMES}, got {workbook_formula.sheetnames}"
        )

    resolved_concepts: dict[str, Any] = {}
    for check in WORKBOOK_CONCEPT_CHECKS:
        sheet_name = check["sheet"]
        if sheet_name not in workbook_formula.sheetnames:
            errors.append(f"Missing required sheet: {sheet_name}")
            continue

        formula_sheet = workbook_formula[sheet_name]
        value_sheet = workbook_values[sheet_name]
        label_cell = _find_label_cell(formula_sheet, check["aliases"])
        if label_cell is None:
            errors.append(
                f"Could not find required concept label on {sheet_name}: {check['aliases']}"
            )
            continue

        value_cells = _find_value_cell(formula_sheet, value_sheet, label_cell)
        if value_cells is None:
            errors.append(
                f"Could not find adjacent numeric/formula cell for {sheet_name} label "
                f"{label_cell.coordinate}={label_cell.value!r}"
            )
            continue

        formula_cell, value_cell = value_cells
        actual_value = value_cell.value

        expected_value = check.get("expected_value")
        expected_abs_value = check.get("expected_abs_value")
        if expected_value is not None:
            ev_list = expected_value if isinstance(expected_value, list) else [expected_value]
            if not any(_close_enough(actual_value, ev) for ev in ev_list):
                errors.append(
                    f"Workbook concept {check['check_id']} on {sheet_name} expected "
                    f"{expected_value}, got {actual_value!r} at {value_cell.coordinate}"
                )
        if expected_abs_value is not None:
            eav_list = expected_abs_value if isinstance(expected_abs_value, list) else [expected_abs_value]
            if not any(_close_enough_abs(actual_value, eav) for eav in eav_list):
                errors.append(
                    f"Workbook concept {check['check_id']} on {sheet_name} expected magnitude "
                    f"{expected_abs_value}, got {actual_value!r} at {value_cell.coordinate}"
                )

        if check.get("require_formula"):
            if not isinstance(formula_cell.value, str) or not formula_cell.value.startswith("="):
                errors.append(
                    f"Workbook concept {check['check_id']} on {sheet_name} must be formula-driven; "
                    f"found {formula_cell.value!r} at {formula_cell.coordinate}"
                )
            else:
                ref_count = _count_formula_refs(formula_cell.value)
                min_refs = int(check.get("min_formula_refs", 1))
                if ref_count < min_refs:
                    errors.append(
                        f"Workbook concept {check['check_id']} on {sheet_name} needs a real "
                        f"formula with at least {min_refs} cell references; found "
                        f"{formula_cell.value!r} at {formula_cell.coordinate}"
                    )

        concept_name = check["concept"]
        if actual_value is not None:
            absolute = expected_abs_value is not None
            if concept_name in resolved_concepts:
                if not _concept_values_consistent(
                    resolved_concepts[concept_name],
                    actual_value,
                    absolute=absolute,
                ):
                    errors.append(
                        f"Inconsistent workbook values for concept {concept_name!r}: "
                        f"{resolved_concepts[concept_name]!r} vs {actual_value!r}"
                    )
            else:
                resolved_concepts[concept_name] = actual_value

    return errors, resolved_concepts, workbook_formula


def _check_pdf(
    pdf_bytes: bytes,
    workbook_concepts: dict[str, Any],
) -> tuple[list[str], dict[str, str]]:
    errors: list[str] = []
    pdf_values = _extract_pdf_values(pdf_bytes)

    for suffix, expected_text in PDF_FIELD_EXPECTATIONS.items():
        actual_text = _lookup_pdf_value(pdf_values, suffix)
        if isinstance(expected_text, tuple):
            if actual_text not in expected_text:
                errors.append(
                    f"PDF field {suffix} mismatch: expected one of {expected_text!r}, got {actual_text!r}"
                )
        elif actual_text != expected_text:
            errors.append(
                f"PDF field {suffix} mismatch: expected {expected_text!r}, got {actual_text!r}"
            )

    for check in PDF_RECONCILIATION_CHECKS:
        concept_name = check["concept"]
        formatter = PDF_FORMATTERS[check["formatter"]]
        workbook_value = workbook_concepts.get(concept_name)
        try:
            expected_text = formatter(workbook_value)
        except Exception:
            errors.append(
                f"Could not derive PDF expectation from workbook concept {concept_name!r}: "
                f"{workbook_value!r}"
            )
            continue

        for suffix in check["suffixes"]:
            actual_text = _lookup_pdf_value(pdf_values, suffix)
            if actual_text != expected_text:
                errors.append(
                    f"PDF field {suffix} mismatch: expected {expected_text!r}, got {actual_text!r}"
                )

    return errors, pdf_values


def score_output_artifacts(*, workbook_bytes: bytes, pdf_bytes: bytes) -> dict[str, Any]:
    workbook_errors, workbook_concepts, workbook_formula = _check_workbook(workbook_bytes)
    pdf_errors, pdf_values = _check_pdf(pdf_bytes, workbook_concepts)
    errors = workbook_errors + pdf_errors
    return {
        "score": 1.0 if not errors else 0.0,
        "passed": not errors,
        "errors": errors,
        "resolved_workbook_concepts": workbook_concepts,
        "workbook_sheetnames": workbook_formula.sheetnames,
        "pdf_nonempty_field_count": len(pdf_values),
    }


def _resolve_paths(args: argparse.Namespace) -> tuple[Path, Path]:
    if args.output_dir is not None:
        return (
            args.output_dir / "tax_workpapers.xlsx",
            args.output_dir / "completed_1040.pdf",
        )
    if args.workbook is not None and args.pdf is not None:
        return args.workbook, args.pdf
    raise ValueError("Provide either --output-dir or both --workbook and --pdf.")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Score federal_income_tax_workpapers output artifacts."
    )
    parser.add_argument("--output-dir", type=Path)
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--pdf", type=Path)
    args = parser.parse_args()

    workbook_path, pdf_path = _resolve_paths(args)
    result = score_output_artifacts(
        workbook_bytes=workbook_path.read_bytes(),
        pdf_bytes=pdf_path.read_bytes(),
    )
    print(json.dumps(result, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
