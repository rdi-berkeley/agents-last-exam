from __future__ import annotations

import argparse
import json
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

EXPECTED_HEADER = (
    "element_name",
    "FIG_code",
    "credited_value",
    "meets_minimum_standard",
)
EXPECTED_ROW_COUNT = 8
MIN_D_SCORE = Decimal("7.0")
MAX_D_SCORE = Decimal("7.2")


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _normalize_text(value: Any) -> str:
    return "" if _is_blank(value) else str(value).strip()


def _normalize_credit(value: Any) -> Decimal:
    if _is_blank(value):
        raise ValueError("credited_value is blank")
    try:
        normalized = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"invalid credited_value: {value!r}") from exc
    return normalized.quantize(Decimal("0.1"))


def _load_rows(blob: bytes) -> list[list[Any]]:
    wb = load_workbook(BytesIO(blob), data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    while rows and all(_is_blank(cell) for cell in rows[-1]):
        rows.pop()
    return rows


def _parse_workbook(blob: bytes) -> dict[str, Any]:
    rows = _load_rows(blob)
    if not rows:
        return {"ok": False, "reason": "workbook is empty"}

    header = tuple("" if cell is None else str(cell).strip() for cell in rows[0])
    if header != EXPECTED_HEADER:
        return {
            "ok": False,
            "reason": "header mismatch",
            "expected_header": list(EXPECTED_HEADER),
            "actual_header": list(header),
        }

    data_rows = rows[1:]
    nonempty_rows = [row for row in data_rows if any(not _is_blank(cell) for cell in row)]
    if len(nonempty_rows) != EXPECTED_ROW_COUNT:
        return {
            "ok": False,
            "reason": "row_count_mismatch",
            "expected_rows": EXPECTED_ROW_COUNT,
            "actual_rows": len(nonempty_rows),
        }

    parsed_rows = []
    credited_total = Decimal("0.0")
    for idx, row in enumerate(nonempty_rows, start=1):
        padded = list(row[: len(EXPECTED_HEADER)])
        if len(padded) < len(EXPECTED_HEADER):
            padded.extend([None] * (len(EXPECTED_HEADER) - len(padded)))
        try:
            credited_value = _normalize_credit(padded[2])
        except ValueError as exc:
            return {"ok": False, "reason": str(exc), "row_index": idx}
        parsed_rows.append(
            {
                "element_name": _normalize_text(padded[0]),
                "FIG_code": _normalize_text(padded[1]),
                "credited_value": credited_value,
                "meets_minimum_standard": _normalize_text(padded[3]),
            }
        )
        credited_total += credited_value

    return {
        "ok": True,
        "rows": parsed_rows,
        "credited_total": credited_total,
    }


def score_workbook_bytes(agent_blob: bytes, reference_blob: bytes) -> dict[str, Any]:
    agent = _parse_workbook(agent_blob)
    if not agent["ok"]:
        return {
            "score": 0.0,
            "reason": agent["reason"],
            **{k: v for k, v in agent.items() if k != "ok"},
        }

    reference = _parse_workbook(reference_blob)
    if not reference["ok"]:
        raise RuntimeError(f"reference workbook invalid: {reference}")

    credited_total = agent["credited_total"]
    if credited_total < MIN_D_SCORE or credited_total > MAX_D_SCORE:
        return {
            "score": 0.0,
            "reason": "d_score_out_of_range",
            "credited_total": float(credited_total),
        }

    agent_pairs = [(row["FIG_code"], row["credited_value"]) for row in agent["rows"]]
    reference_pairs = [(row["FIG_code"], row["credited_value"]) for row in reference["rows"]]
    if agent_pairs != reference_pairs:
        mismatch_index = next(
            idx
            for idx, (agent_row, ref_row) in enumerate(zip(agent_pairs, reference_pairs), start=1)
            if agent_row != ref_row
        )
        return {
            "score": 0.0,
            "reason": "ordered_match_failed",
            "mismatch_index": mismatch_index,
            "agent_row": [agent_pairs[mismatch_index - 1][0], float(agent_pairs[mismatch_index - 1][1])],
            "reference_row": [
                reference_pairs[mismatch_index - 1][0],
                float(reference_pairs[mismatch_index - 1][1]),
            ],
            "credited_total": float(credited_total),
        }

    return {
        "score": 1.0,
        "reason": "pass",
        "credited_total": float(credited_total),
        "row_count": len(agent_pairs),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Verify aerobics difficulty workbook output.")
    parser.add_argument("--agent", required=True, help="Path to the agent workbook")
    parser.add_argument("--reference", required=True, help="Path to the hidden reference workbook")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    report = score_workbook_bytes(Path(args.agent).read_bytes(), Path(args.reference).read_bytes())
    print(json.dumps(report))


if __name__ == "__main__":
    main()
