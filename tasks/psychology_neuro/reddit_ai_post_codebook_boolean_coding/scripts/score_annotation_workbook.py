#!/usr/bin/env python
"""Local scorer for the Reddit AI post codebook boolean-coding task."""

import argparse
import json
from io import BytesIO

from openpyxl import load_workbook


EXPECTED_COLUMN_COUNT = 44
FIRST_DATA_ROW = 2
LAST_DATA_ROW = 96
FIRST_SCORED_COLUMN = 6   # F
LAST_SCORED_COLUMN = 44   # AR


def _normalize_bool(value):
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        if value == 1:
            return 1
        if value == 0:
            return 0
    if isinstance(value, str):
        cleaned = value.strip().lower()
        if cleaned in {"", "0", "false", "f", "no"}:
            return 0
        if cleaned in {"1", "true", "t", "yes"}:
            return 1
    raise ValueError(f"invalid boolean-like cell value: {value!r}")


def _load_first_sheet(workbook_bytes: bytes, *, label: str):
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
    if not workbook.sheetnames:
        raise ValueError(f"{label} workbook has no sheets")
    return workbook[workbook.sheetnames[0]]


def _header_values(sheet):
    return [sheet.cell(row=1, column=col_idx).value for col_idx in range(1, EXPECTED_COLUMN_COUNT + 1)]


def score_workbooks(candidate_bytes: bytes, reference_bytes: bytes) -> dict[str, object]:
    try:
        candidate_sheet = _load_first_sheet(candidate_bytes, label="candidate")
        reference_sheet = _load_first_sheet(reference_bytes, label="reference")
    except Exception as exc:
        return {"score": 0.0, "hard_fail": "unreadable_workbook", "error": str(exc)}

    try:
        candidate_headers = _header_values(candidate_sheet)
        reference_headers = _header_values(reference_sheet)
        if candidate_sheet.max_column < EXPECTED_COLUMN_COUNT or reference_sheet.max_column < EXPECTED_COLUMN_COUNT:
            return {
                "score": 0.0,
                "hard_fail": "wrong_header_width",
                "error": f"expected {EXPECTED_COLUMN_COUNT} columns",
            }
        if candidate_headers != reference_headers:
            return {"score": 0.0, "hard_fail": "header_mismatch"}

        for row_idx in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            for col_idx in range(1, 6):
                if candidate_sheet.cell(row=row_idx, column=col_idx).value != reference_sheet.cell(
                    row=row_idx, column=col_idx
                ).value:
                    return {
                        "score": 0.0,
                        "hard_fail": "metadata_mismatch",
                        "row": row_idx,
                        "column": col_idx,
                    }

            focus_sum = 0
            for col_idx in range(6, 9):
                focus_sum += _normalize_bool(candidate_sheet.cell(row=row_idx, column=col_idx).value)
            if focus_sum != 1:
                return {
                    "score": 0.0,
                    "hard_fail": "invalid_focus_section",
                    "row": row_idx,
                    "focus_positive_count": focus_sum,
                }
    except Exception as exc:
        return {"score": 0.0, "hard_fail": "invalid_structure", "error": str(exc)}

    wrong_cells = 0
    total_cells = 0
    try:
        for row_idx in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            for col_idx in range(FIRST_SCORED_COLUMN, LAST_SCORED_COLUMN + 1):
                total_cells += 1
                candidate_value = _normalize_bool(candidate_sheet.cell(row=row_idx, column=col_idx).value)
                reference_value = _normalize_bool(reference_sheet.cell(row=row_idx, column=col_idx).value)
                if candidate_value != reference_value:
                    wrong_cells += 1
    except Exception as exc:
        return {"score": 0.0, "hard_fail": "invalid_scored_cell", "error": str(exc)}

    score = 1.0 - (wrong_cells / total_cells)
    return {
        "score": score,
        "wrong_cells": wrong_cells,
        "total_cells": total_cells,
        "accuracy": score,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--candidate", required=True)
    parser.add_argument("--reference", required=True)
    args = parser.parse_args()

    with open(args.candidate, "rb") as candidate_file:
        candidate_bytes = candidate_file.read()
    with open(args.reference, "rb") as reference_file:
        reference_bytes = reference_file.read()

    print(json.dumps(score_workbooks(candidate_bytes, reference_bytes), indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
