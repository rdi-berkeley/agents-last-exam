"""Workbook scorer for social_media_health_extraction_instance_1."""

from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

REQUIRED_COLUMNS = [
    "query_id",
    "source",
    "post_id",
    "symptom",
    "social_factor",
    "sentiment",
    "evidence_text",
]
KEY_FIELDS = ("query_id", "source", "post_id")
FIELD_WEIGHTS = {
    "symptom": 0.25,
    "social_factor": 0.25,
    "sentiment": 0.20,
    "evidence_text": 0.30,
}
SCHEMA_FIELDS = ("symptom", "social_factor", "sentiment")


@dataclass(frozen=True)
class ScoreResult:
    score: float
    passed: bool
    reason: str
    hard_gate: str | None
    row_count_agent: int
    row_count_reference: int
    component_scores: dict[str, float]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def _load_schema(schema_json_text: str) -> dict[str, set[str]]:
    raw = json.loads(schema_json_text)
    if not isinstance(raw, dict):
        raise ValueError("schema must be a JSON object")

    schema: dict[str, set[str]] = {}
    for field in SCHEMA_FIELDS:
        values = raw.get(field)
        if not isinstance(values, list):
            raise ValueError(f"schema field {field} must be a list")
        normalized: set[str] = set()
        for value in values:
            if not isinstance(value, str):
                raise ValueError(f"schema field {field} contains a non-string member")
            cleaned = value.strip()
            if not cleaned:
                raise ValueError(f"schema field {field} contains an empty member")
            normalized.add(cleaned)
        if not normalized:
            raise ValueError(f"schema field {field} must be non-empty")
        schema[field] = normalized
    return schema


def _load_rows(payload: bytes) -> tuple[list[str], list[dict[str, str]]]:
    workbook = load_workbook(BytesIO(payload), data_only=True)
    if len(workbook.worksheets) != 1:
        raise ValueError("workbook must contain exactly one sheet")
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("empty workbook")

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    parsed_rows: list[dict[str, str]] = []
    for row in rows[1:]:
        if row is None:
            continue
        values = ["" if cell is None else str(cell).strip() for cell in row[: len(header)]]
        if not any(values):
            continue
        parsed_rows.append(dict(zip(header, values, strict=False)))
    return header, parsed_rows


def _key_for_row(row: dict[str, str]) -> tuple[str, str, str]:
    return tuple(row.get(field, "").strip() for field in KEY_FIELDS)  # type: ignore[return-value]


def _validate_rows(rows: list[dict[str, str]], schema: dict[str, set[str]]) -> str | None:
    seen: set[tuple[str, str, str]] = set()
    for row in rows:
        key = _key_for_row(row)
        if any(not part for part in key):
            return "missing_key_field"
        if key in seen:
            return "duplicate_key"
        seen.add(key)
        for field in ("symptom", "social_factor", "sentiment"):
            value = row.get(field, "").strip()
            if value not in schema[field]:
                return f"out_of_schema_{field}"
    return None


def _evidence_match(agent: str, reference: str) -> float:
    agent_norm = _normalize_text(agent)
    ref_norm = _normalize_text(reference)
    if not agent_norm or not ref_norm:
        return 0.0
    return 1.0 if agent_norm in ref_norm or ref_norm in agent_norm else 0.0


def score_workbook_pair(
    *,
    agent_workbook_bytes: bytes,
    reference_workbook_bytes: bytes,
    schema_json_text: str,
    threshold: float,
) -> ScoreResult:
    try:
        agent_header, agent_rows = _load_rows(agent_workbook_bytes)
    except Exception:
        return ScoreResult(0.0, False, "hard_gate_failure", "invalid_agent_workbook", 0, 0, {})
    try:
        reference_header, reference_rows = _load_rows(reference_workbook_bytes)
    except Exception:
        return ScoreResult(0.0, False, "reference_error", "invalid_reference_workbook", 0, 0, {})
    try:
        schema = _load_schema(schema_json_text)
    except Exception:
        return ScoreResult(
            0.0,
            False,
            "reference_error",
            "invalid_reference_schema",
            len(agent_rows),
            len(reference_rows),
            {},
        )

    if agent_header != REQUIRED_COLUMNS:
        return ScoreResult(
            0.0,
            False,
            "hard_gate_failure",
            "agent_header_mismatch",
            len(agent_rows),
            len(reference_rows),
            {},
        )
    if reference_header != REQUIRED_COLUMNS:
        return ScoreResult(
            0.0,
            False,
            "reference_error",
            "reference_header_mismatch",
            len(agent_rows),
            len(reference_rows),
            {},
        )

    agent_error = _validate_rows(agent_rows, schema)
    if agent_error:
        return ScoreResult(
            0.0,
            False,
            "hard_gate_failure",
            agent_error,
            len(agent_rows),
            len(reference_rows),
            {},
        )

    reference_error = _validate_rows(reference_rows, schema)
    if reference_error:
        return ScoreResult(
            0.0,
            False,
            "reference_error",
            reference_error,
            len(agent_rows),
            len(reference_rows),
            {},
        )

    agent_map = {_key_for_row(row): row for row in agent_rows}
    reference_map = {_key_for_row(row): row for row in reference_rows}
    all_keys = sorted(set(agent_map) | set(reference_map))
    if not all_keys:
        return ScoreResult(0.0, False, "hard_gate_failure", "no_rows", 0, 0, {})

    total = 0.0
    component_totals = {field: 0.0 for field in FIELD_WEIGHTS}
    for key in all_keys:
        agent_row = agent_map.get(key, {})
        reference_row = reference_map.get(key, {})
        row_score = 0.0
        for field, weight in FIELD_WEIGHTS.items():
            if field == "evidence_text":
                component = _evidence_match(
                    agent_row.get(field, ""),
                    reference_row.get(field, ""),
                )
            else:
                component = 1.0 if agent_row.get(field, "") == reference_row.get(field, "") else 0.0
            component_totals[field] += component
            row_score += component * weight
        total += row_score

    score = total / len(all_keys)
    component_scores = {field: value / len(all_keys) for field, value in component_totals.items()}
    passed = score >= threshold
    return ScoreResult(
        score=score,
        passed=passed,
        reason="ok" if passed else "below_threshold",
        hard_gate=None,
        row_count_agent=len(agent_rows),
        row_count_reference=len(reference_rows),
        component_scores=component_scores,
    )
